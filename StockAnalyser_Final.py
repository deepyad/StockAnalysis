# https://www.perplexity.ai/search/just-add-eps-earning-per-share-ParpkhPlRNOhMCdn00iAqQ

import pandas as pd
import yfinance as yf
import os
import datetime
import numpy as np
import time
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend for Windows
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# ================== RSI, MACD, PE Calculation Functions ==================

import requests
from bs4 import BeautifulSoup

# def get_latest_news_link(ticker_symbol, positive=True):
#     """
#     Finds the most recent positive or negative news link for a given ticker.
#     If positive is True, searches for the most positive; otherwise, most negative.
#     Returns the news title and its direct link, or None if not found.
#     """
#     # We'll use Google News search for the ticker
#     # NB: This is for demonstration. In production, use a real news API for reliability.
#     query = f"{ticker_symbol} stock"
#     url = f"https://news.google.com/search?q={query}&hl=en-IN&gl=IN&ceid=IN:en"
#     headers = {"User-Agent": "Mozilla/5.0"}
#     try:
#         resp = requests.get(url, headers=headers)
#         soup = BeautifulSoup(resp.text, 'html.parser')
#         articles = soup.find_all('article')
#         news_items = []
#         for article in articles:
#             header = article.find('h3')
#             if not header:
#                 continue
#             title = header.text.lower()
#             link_tag = header.find('a')
#             if not link_tag:
#                 continue
#             link = link_tag['href']
#             # Google News uses relative URLs for articles
#             if not link.startswith('http'):
#                 link = 'https://news.google.com' + link[1:]
#             news_items.append((title, link))
#         # Define simple lists of positive/negative keywords
#         positives = ['surge', 'profit', 'beats', 'rise', 'growth', 'record', 'gain', 'buy', 'up', 'upgrade', 'outperform']
#         negatives = ['loss', 'decline', 'drops', 'down', 'plunge', 'sell', 'downgrade', 'bearish', 'fall', 'lower', 'misses']
#         for keywords in (positives if positive else negatives):
#             for title, link in news_items:
#                 if keywords in title:
#                     return title.title(), link
#         # If not found, return the first latest news headline
#         if news_items:
#             return news_items[0].title(), news_items[12]
#     except Exception as e:
#         print(f"Failed to get news for {ticker_symbol}: {e}")
#     return None, None
import requests
from bs4 import BeautifulSoup

import requests

def get_latest_news_link_newsdata(ticker_symbol, positive=True, apikey="pub_0ddc8208bc3746b2b1470feaf234c927", days_back=7):
    """
    Fetches the latest news from the past N days for the given ticker using NewsData.io API.
    Returns headline and url.
    
    IMPORTANT: To get news working, you need a valid NewsData.io API key.
    Get one from: https://newsdata.io/api
    Then replace the 'apikey' parameter value above.
    """
    # Remove .NS suffix if present for better search results
    query = ticker_symbol.replace('.NS', '').replace('.BO', '')
    url = f"https://newsdata.io/api/1/news"
    
    # Calculate date range (past N days)
    from datetime import datetime, timedelta
    date_from = (datetime.now() - timedelta(days=days_back)).strftime('%Y-%m-%d')
    
    # Note: NewsData.io API parameters
    # The "timeframe" parameter is not supported in API v1
    # API automatically returns recent news
    params = {
        "apikey": apikey,
        "q": query,
        "language": "en",
        "country": "in",
        "category": "business"
    }
    try:
        resp = requests.get(url, params=params, timeout=10)
        
        # Check response status
        if resp.status_code != 200:
            print(f"  ⚠ News API returned status {resp.status_code} for {ticker_symbol}")
            return None, None
            
        data = resp.json()
        
        # Check for API error messages
        if 'status' in data and data['status'] == 'error':
            error_msg = data.get('message', 'Unknown error')
            print(f"  ⚠ News API error for {ticker_symbol}: {error_msg}")
            return None, None
        
        # Check if response is valid and has results
        if 'results' not in data:
            return None, None
            
        # Ensure results is a list, not a string
        if not isinstance(data.get('results'), list):
            print(f"  ⚠ Unexpected results format for {ticker_symbol}")
            return None, None
            
        # If no results, return None
        if not data.get('results') or len(data.get('results', [])) == 0:
            # Only print for first few stocks to avoid spam
            import random
            if random.random() < 0.1:  # 10% chance to print
                print(f"  ℹ No news results for {ticker_symbol}")
            return None, None
            
        keywords_pos = ['surge', 'profit', 'beats', 'rise', 'growth', 'record', 'gain', 'buy', 'up', 'upgrade', 'outperform', 
                       'earnings', 'results', 'quarter', 'sales', 'revenue', 'expansion', 'deal', 'contract', 'order']
        keywords_neg = ['loss', 'decline', 'drops', 'down', 'plunge', 'sell', 'downgrade', 'bearish', 'fall', 'lower', 'misses',
                       'concern', 'issue', 'warn', 'delay', 'cut', 'layoff']
        
        # First, try to find news with relevant keywords
        search_keywords = keywords_pos if positive else keywords_neg
        for news in data.get('results', []):
            title = news.get('title', '').lower()
            description = news.get('description', '').lower()
            link = news.get('link', None)
            if any(k in title or k in description for k in search_keywords) and link:
                return news.get('title'), link
        
        # fallback: return latest headline (if exists)
        if data.get('results') and len(data.get('results', [])) > 0:
            return data['results'][0].get('title', None), data['results'][0].get('link', None)
    except Exception as e:
        # Print error for debugging
        print(f"  ⚠ News API error for {ticker_symbol}: {e}")
        return None, None

# Usage:
# title, link = get_latest_news_link_newsdata("ASHOKLEY.NS", positive=True, apikey="YOUR_API_KEY")
# print(title, link)


# Usage:
# title, link = get_latest_news_link("ASHOKLEY.NS", positive=True)
# print("Most positive:", title, link)

# Example Usage:
# positive_news_title, positive_news_link = get_latest_news_link("ASHOKLEY.NS", positive=True)
# print("Most positive:", positive_news_title, positive_news_link)
# negative_news_title, negative_news_link = get_latest_news_link("ASHOKLEY.NS", positive=False)
# print("Most negative:", negative_news_title, negative_news_link)


def calculate_RSI(prices, period=14):
    delta = prices.diff()
    gain = np.where(delta > 0, delta, 0)
    loss = np.where(delta < 0, -delta, 0)
    avg_gain = pd.Series(gain).rolling(window=period).mean()
    avg_loss = pd.Series(loss).rolling(window=period).mean()
    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    return rsi

def classify_RSI_per_day(rsi_series):
    status = rsi_series.apply(lambda x: "Overbought" if x > 70 else ("Oversold" if x < 30 else "Neutral"))
    return status

def calculate_MACD(prices, short_window=12, long_window=26, signal_window=9):
    ema_short = prices.ewm(span=short_window, adjust=False).mean()
    ema_long = prices.ewm(span=long_window, adjust=False).mean()
    macd_line = ema_short - ema_long
    signal_line = macd_line.ewm(span=signal_window, adjust=False).mean()
    return macd_line, signal_line

# def classify_MACD_per_day(macd_series, signal_series, threshold=0.7):
#     diff = macd_series - signal_series
#     def classify_val(x):
#         if x > threshold:
#             return "Overbought"
#         # elif x < -threshold:
#         elif x < .3:
#             return "Oversold"
#         else:
#             return "Neutral"
#     status = diff.apply(classify_val)
#     return status
def classify_MACD_per_day(macd_series, signal_series):
    """
    Classifies MACD status with extreme/medium/neutral levels.
    Extreme Overbought: diff > 1.5
    Medium Overbought: diff > 0.7
    Neutral: -0.7 <= diff <= 0.7
    Medium Oversold: diff < -0.7
    Extreme Oversold: diff < -1.5
    """
    diff = macd_series - signal_series

    def classify_val(x):
        if x > 1.5:
            return "Extreme Overbought"
        elif x > 0.7:
            return "Medium Overbought"
        elif x < -1.5:
            return "Extreme Oversold"
        elif x < -0.7:
            return "Medium Oversold"
        else:
            return "Neutral"

    status = diff.apply(classify_val)
    return status

def get_PE_history(ticker, period="5y", interval="1mo"):
    # Get monthly close prices for 5 years
    stock = yf.Ticker(ticker)
    hist = stock.history(period=period, interval=interval)
    try:
        info = stock.info
        trailing_eps = info.get('trailingEps', None)
        # If EPS is not available, return empty series
        if trailing_eps in [None, 0, 'nan']:
            return pd.Series([np.nan]*len(hist), index=hist.index)
        pe_series = hist['Close'] / trailing_eps
        return pe_series
    except Exception as e:
        print(f"Error fetching PE history for {ticker}: {e}")
        return pd.Series([np.nan]*len(hist), index=hist.index)

def get_latest_PE(ticker_symbol):
    try:
        stock = yf.Ticker(ticker_symbol)
        info = stock.info
        trailing_eps = info.get('trailingEps', None)
        latest_price = info.get('regularMarketPrice', None)
        if trailing_eps in [None, 0, 'nan'] or latest_price in [None, 'nan']:
            return np.nan
        return latest_price / trailing_eps
    except Exception as e:
        print(f"Error fetching PE for {ticker_symbol}: {e}")
        return np.nan

# ================== Utility Functions ==================

def format_percent_change(change):
    if pd.isna(change):
        return None
    return f"{'+' if change > 0 else ''}{change:.2f}%"

def get_analyst_recommendation(stock, ticker_symbol):
    try:
        recommendations = stock.recommendations
        if recommendations is not None and not recommendations.empty:
            latest_rec = recommendations.iloc[-1]
            if 'To Grade' in latest_rec:
                return latest_rec['To Grade'].upper()
        info = stock.info
        if 'recommendationKey' in info:
            rec_key = info['recommendationKey'].upper()
            rec_map = {
                'STRONG_BUY': 'STRONG BUY',
                'BUY': 'BUY',
                'HOLD': 'HOLD',
                'UNDERPERFORM': 'SELL',
                'SELL': 'STRONG SELL'
            }
            return rec_map.get(rec_key, rec_key)
        return "N/A"
    except:
        return "N/A"

def get_price_target_pct(stock, current_price, ticker_symbol):
    try:
        info = stock.info
        if 'targetMeanPrice' in info and info['targetMeanPrice'] is not None:
            target_price = info['targetMeanPrice']
            return ((target_price - current_price) / current_price) * 100
        return np.nan
    except:
        return np.nan

def generate_recommendation(week_change, month_change, three_month_change, analyst_rec, target_pct):
    if analyst_rec not in ["N/A"]:
        return analyst_rec
    week_change = 0 if pd.isna(week_change) else week_change
    month_change = 0 if pd.isna(month_change) else month_change
    three_month_change = 0 if pd.isna(three_month_change) else three_month_change
    target_pct = 0 if pd.isna(target_pct) else target_pct
    if target_pct != 0:
        score = (0.15 * week_change) + (0.2 * month_change) + (0.25 * three_month_change) + (0.4 * target_pct)
    else:
        score = (0.2 * week_change) + (0.3 * month_change) + (0.5 * three_month_change)
    if score > 15:
        return "BUY"
    elif score < -10:
        return "SELL"
    else:
        return "HOLD"

# ================== Chart Creation Function ==================
def create_indicator_chart(ticker, hist, rsi_series, macd_line, signal_line, pe_series, save_dir="charts"):
    """Create indicator charts for MACD, RSI, and PE ratio"""
    try:
        os.makedirs(save_dir, exist_ok=True)
        fig, axes = plt.subplots(3, 1, figsize=(4, 5), sharex=False,
                                 gridspec_kw={'height_ratios': [2, 1, 1]})
        # MACD graph (no close price)
        axes[0].plot(hist.index, macd_line, label='MACD', color='blue', linewidth=1)
        axes[0].plot(hist.index, signal_line, label='Signal', color='red', linewidth=1, linestyle='--')
        axes[0].legend(fontsize=6)
        axes[0].set_title(f"{ticker} - MACD", fontsize=8)
        # RSI graph
        axes[1].plot(hist.index, rsi_series, label='RSI', color='purple', linewidth=1)
        axes[1].axhline(70, color='red', linestyle='--', linewidth=0.7)
        axes[1].axhline(30, color='green', linestyle='--', linewidth=0.7)
        axes[1].set_ylim(0, 100)
        axes[1].legend(fontsize=6)
        axes[1].set_title("RSI", fontsize=8)
        # PE Ratio graph (5-year monthly)
        if pe_series is not None and not pe_series.empty:
            axes[2].plot(pe_series.index, pe_series, label='PE Ratio', color='orange', linewidth=1)
            axes[2].legend(fontsize=6)
            axes[2].set_title("PE Ratio (5yr)", fontsize=8)
        else:
            axes[2].set_title("PE Ratio unavailable", fontsize=8)
        plt.tight_layout()
        filepath = os.path.join(save_dir, f"{ticker}_chart.png")
        plt.savefig(filepath, dpi=80, bbox_inches='tight')
        plt.close(fig)
        return filepath
    except Exception as e:
        print(f"Error creating chart for {ticker}: {e}")
        return None

# ================== Main Data Function ==================
def get_stock_data(ticker_symbol):
    try:
        stock = yf.Ticker(ticker_symbol)
        hist = stock.history(period="100d")
        if len(hist) < 2:
            empty_series = pd.Series(dtype=np.float64)
            empty_status = pd.Series(dtype=object)
            return np.nan, np.nan, np.nan, "N/A", np.nan, empty_series, empty_status, empty_series, empty_series, empty_status, np.nan, empty_series, "", None
        latest_close = hist['Close'].iloc[-1]
        latest_date = hist.index[-1].date()
        hist_dates = [d.date() for d in hist.index]
        def find_closest_trading_day(target_date):
            valid_dates = [d for d in hist_dates if d <= target_date]
            if not valid_dates:
                return 0
            closest_date = max(valid_dates)
            return hist_dates.index(closest_date)
        week_idx = find_closest_trading_day(latest_date - datetime.timedelta(days=7))
        month_idx = find_closest_trading_day(latest_date - datetime.timedelta(days=30))
        three_month_idx = find_closest_trading_day(latest_date - datetime.timedelta(days=90))
        week_chg = ((latest_close - hist['Close'].iloc[week_idx]) / hist['Close'].iloc[week_idx]) * 100
        month_chg = ((latest_close - hist['Close'].iloc[month_idx]) / hist['Close'].iloc[month_idx]) * 100
        three_m_chg = ((latest_close - hist['Close'].iloc[three_month_idx]) / hist['Close'].iloc[three_month_idx]) * 100
        recommendation = get_analyst_recommendation(stock, ticker_symbol)
        target_pct = get_price_target_pct(stock, latest_close, ticker_symbol)
        rsi_series = calculate_RSI(hist['Close'])
        rsi_status = classify_RSI_per_day(rsi_series)
        macd_line, signal_line = calculate_MACD(hist['Close'])
        macd_status = classify_MACD_per_day(macd_line, signal_line)
        # PE ratio values
        pe_series = get_PE_history(ticker_symbol, period="5y", interval="1mo")
        latest_pe = get_latest_PE(ticker_symbol)
        
        # Create chart - handle errors gracefully
        try:
            chart_path = create_indicator_chart(ticker_symbol, hist, rsi_series, macd_line, signal_line, pe_series)
            if chart_path is None:
                chart_path = ""  # Fallback if chart creation failed
        except Exception as e:
            print(f"Warning: Could not create chart for {ticker_symbol}: {e}")
            chart_path = ""
        
        # Get latest news
        try:
            latest_news = get_latest_news_link_newsdata(ticker_symbol)
        except Exception as e:
            print(f"Warning: Could not fetch news for {ticker_symbol}: {e}")
            latest_news = (None, None)
        
        return (week_chg, month_chg, three_m_chg, recommendation, target_pct,
                rsi_series, rsi_status, macd_line, signal_line, macd_status, latest_pe, pe_series, chart_path, latest_news)
    except Exception as e:
        print(f"Error fetching data for {ticker_symbol}: {e}")
        empty_series = pd.Series(dtype=np.float64)
        empty_status = pd.Series(dtype=object)
        return np.nan, np.nan, np.nan, "N/A", np.nan, empty_series, empty_status, empty_series, empty_series, empty_status, np.nan, empty_series, "", None

# ================ News Polarity Scoring ================
def calculate_news_polarity(headline):
    """
    Calculate sentiment score based on news headline
    Returns: +1.0 (positive), -1.0 (negative), 0.0 (neutral), or None if no news
    """
    if not headline or headline == "No recent news" or pd.isna(headline):
        return None
    
    headline_lower = str(headline).lower()
    
    # Positive keywords (strong sentiment)
    strong_positive = ['surge', 'soars', 'boom', 'rocket', 'explosive', 'breakthrough', 
                       'record high', 'all-time high', 'multi-bagger', 'skyrocket']
    
    # Positive keywords (moderate)
    positive = ['profit', 'beats', 'rise', 'growth', 'record', 'gain', 'buy', 
                'up', 'upgrade', 'outperform', 'earnings beat', 'revenue growth',
                'expansion', 'deal', 'contract', 'order', 'partnership', 'acquisition']
    
    # Negative keywords (strong sentiment)
    strong_negative = ['plunge', 'collapse', 'crash', 'disaster', 'crisis', 'bankruptcy',
                       'warning', 'cut', 'exit', 'shut down']
    
    # Negative keywords (moderate)
    negative = ['loss', 'decline', 'drops', 'down', 'sell', 'downgrade', 'bearish', 
                'fall', 'lower', 'misses', 'concern', 'issue', 'delay', 'layoff',
                'slump', 'worst', 'weak', 'disappointment']
    
    # Check for strong sentiments first
    if any(word in headline_lower for word in strong_positive):
        return 1.0
    if any(word in headline_lower for word in strong_negative):
        return -1.0
    
    # Count positive vs negative
    pos_count = sum(1 for word in positive if word in headline_lower)
    neg_count = sum(1 for word in negative if word in headline_lower)
    
    if pos_count > neg_count:
        return 0.5
    elif neg_count > pos_count:
        return -0.5
    else:
        return 0.0

# ================ Holdings Processing ================
def process_holdings_file(
    input_file="./holdings_with_changes_14_Aug.csv",
    output_file=None):

    import re
    from datetime import datetime
    
    # Generate output filename with timestamp if not provided
    if output_file is None:
        now = datetime.now()
        output_file = f"OUTPUT_{now.strftime('%d_%b_%Y')}_{now.strftime('%H%M%S')}.xlsx"
    
    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' not found in {os.getcwd()}")
        return
    holdings_df = pd.read_csv(input_file)
    ticker_col = "Instrument"
    holdings_df['Last_Week_Change_%'] = None
    holdings_df['1_Month_Change_%'] = None
    holdings_df['3_Month_Change_%'] = None
    holdings_df['Broker_Recommendation'] = None
    holdings_df['12_Month_Target_%'] = None
    holdings_df['RSI_Score'] = None
    holdings_df['RSI_Status'] = None
    holdings_df['MACD_Value'] = None
    holdings_df['Signal_Value'] = None
    holdings_df['MACD_Status'] = None
    holdings_df['PE_Ratio'] = None
    holdings_df['Indicator_Chart'] = None
    holdings_df['Latest_News_Headline'] = None
    holdings_df['Latest_News_Link'] = None
    holdings_df['News_Polarity_Score'] = None
    holdings_df['Deviation_From_Target_%'] = None
    for i, row in holdings_df.iterrows():
        ticker = str(row[ticker_col]).strip()
        if pd.isna(ticker) or ticker == "":
            continue
        if "." in ticker:
            ticker = ticker.split(".")[0]
        nse_ticker = f"{ticker}.NS"
        print(f"Processing {nse_ticker} ({i+1}/{len(holdings_df)}) ...")
        (week_chg, month_chg, three_m_chg, analyst_rec, target_pct,
         rsi_series, rsi_status, macd_line, sig_line, macd_status,
         latest_pe, pe_series, chart_path,latest_news) = get_stock_data(nse_ticker)
        holdings_df.at[i, 'Last_Week_Change_%'] = format_percent_change(week_chg)
        holdings_df.at[i, '1_Month_Change_%'] = format_percent_change(month_chg)
        holdings_df.at[i, '3_Month_Change_%'] = format_percent_change(three_m_chg)
        holdings_df.at[i, '12_Month_Target_%'] = format_percent_change(target_pct)
        holdings_df.at[i, 'Broker_Recommendation'] = generate_recommendation(week_chg, month_chg, three_m_chg, analyst_rec, target_pct)
        holdings_df.at[i, 'RSI_Score'] = round(rsi_series.iloc[-1], 2) if not rsi_series.empty and pd.notna(rsi_series.iloc[-1]) else None
        holdings_df.at[i, 'RSI_Status'] = rsi_status.iloc[-1] if not rsi_status.empty else None
        holdings_df.at[i, 'MACD_Value'] = round(macd_line.iloc[-1], 4) if not macd_line.empty and pd.notna(macd_line.iloc[-1]) else None
        holdings_df.at[i, 'Signal_Value'] = round(sig_line.iloc[-1], 4) if not sig_line.empty and pd.notna(sig_line.iloc[-1]) else None
        holdings_df.at[i, 'MACD_Status'] = macd_status.iloc[-1] if not macd_status.empty else None
        holdings_df.at[i, 'PE_Ratio'] = round(latest_pe, 2) if pd.notna(latest_pe) else None
        holdings_df.at[i, 'Indicator_Chart'] = chart_path if chart_path else ""
        
        # Print chart creation status
        if chart_path:
            print(f"  ✓ Chart created: {chart_path}")
        else:
            print(f"  ⚠ Chart not created for {nse_ticker}")
        
        # Store news headline and link separately
        if latest_news is not None and len(latest_news) == 2:
            holdings_df.at[i, 'Latest_News_Headline'] = latest_news[0] if latest_news[0] else "No recent news"
            holdings_df.at[i, 'Latest_News_Link'] = latest_news[1] if latest_news[1] else ""
            # Calculate news polarity score
            headline = latest_news[0] if latest_news[0] else "No recent news"
            holdings_df.at[i, 'News_Polarity_Score'] = calculate_news_polarity(headline)
        else:
            holdings_df.at[i, 'Latest_News_Headline'] = "No recent news"
            holdings_df.at[i, 'Latest_News_Link'] = ""
            holdings_df.at[i, 'News_Polarity_Score'] = None

        time.sleep(0.5)

    # Calculate Deviation from Target
    # Deviation shows: How much more/less the stock needs to move to reach its 12-month target
    # Example: Target = +30%, 1-Month Performance = +5% → Deviation = +25% (needs 25% more upside)
    def calculate_deviation_from_target(row):
        """Calculate how far current performance is from target.
        Positive deviation = need more upside to reach target
        Negative deviation = already exceeded target"""
        try:
            target_str = row.get('12_Month_Target_%', '')
            recent_perf_str = row.get('1_Month_Change_%', '') or row.get('Last_Week_Change_%', '')
            
            if pd.isna(target_str) or target_str is None or target_str == "":
                return None
            
            if pd.isna(recent_perf_str) or recent_perf_str is None or recent_perf_str == "":
                return None
            
            # Parse percentage values (remove % and + signs)
            try:
                target = float(str(target_str).replace('%', '').replace('+', '').strip())
                recent_perf = float(str(recent_perf_str).replace('%', '').replace('+', '').strip())
                deviation = target - recent_perf
                return f"{'+' if deviation > 0 else ''}{deviation:.2f}%"
            except:
                return None
        except:
            return None
    
    holdings_df['Deviation_From_Target_%'] = holdings_df.apply(calculate_deviation_from_target, axis=1)

    # 1. Extract numeric 12_Month_Target for sorting
    def parse_pct(val):
        if pd.isna(val):
            return -float('inf')
        try:
            return float(str(val).replace('%','').replace('+',''))
        except Exception:
            return -float('inf')
    holdings_df['_sort_12mo'] = holdings_df['12_Month_Target_%'].apply(parse_pct)

    # 2. Custom sort order for Broker_Recommendation
    rec_sort = {
        "STRONG BUY": 1,
        "BUY": 2,
        "HOLD": 3, "NONE": 3, "N/A": 3,
        "SELL": 4, "STRONG SELL": 4
    }
    def get_rec_rank(val):
        if isinstance(val, str):
            return rec_sort.get(val.strip().upper(), 5)
        return 5
    holdings_df['_rec_rank'] = holdings_df['Broker_Recommendation'].apply(get_rec_rank)

    # 3. Sort by rec_rank, then by _sort_12mo descending
    holdings_df = holdings_df.sort_values(by=['_rec_rank', '_sort_12mo'], ascending=[True, False])
    holdings_df = holdings_df.drop(columns=['_sort_12mo', '_rec_rank'])

    # Write to Excel using the provided output_file parameter
    holdings_df.to_excel(output_file, index=False)
    
    # Format Excel to add hyperlinks for news links and color coding
    try:
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Find column indices for formatting
        header_row = 1
        col_indices = {}
        for col_idx, cell in enumerate(ws[header_row], 1):
            col_indices[cell.value] = col_idx
        
        # Add hyperlinks for news links
        if 'Latest_News_Link' in col_indices:
            col_idx = col_indices['Latest_News_Link']
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and cell.value != "" and cell.value != "No recent news":
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
        
        # Add color coding for P&L column (negative = red, positive = green)
        if 'P&L' in col_indices:
            col_idx = col_indices['P&L']
            from openpyxl.styles import PatternFill, Font
            red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
            green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")  # Light green
            dark_red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Dark red
            
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    # Convert value to float (handle comma separators)
                    val_str = str(cell.value).replace(',', '').replace('₹', '').strip()
                    if val_str and val_str != 'None':
                        val = float(val_str)
                        if val < 0:
                            # Use darker red for larger losses, lighter for smaller
                            if val < -50000:
                                cell.fill = dark_red_fill
                            else:
                                cell.fill = red_fill
                        elif val > 0:
                            cell.fill = green_fill
                except (ValueError, AttributeError):
                    pass
        
        # Add color coding for Net chg. column (negative = red, positive = green)
        if 'Net chg.' in col_indices:
            col_idx = col_indices['Net chg.']
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    val_str = str(cell.value).replace('%', '').replace('+', '').strip()
                    if val_str and val_str != 'None':
                        val = float(val_str)
                        if val < 0:
                            if val < -20:
                                cell.fill = dark_red_fill
                            else:
                                cell.fill = red_fill
                        elif val > 0:
                            cell.fill = green_fill
                except (ValueError, AttributeError):
                    pass
        
        # Add color coding for News_Polarity_Score column
        if 'News_Polarity_Score' in col_indices:
            col_idx = col_indices['News_Polarity_Score']
            strong_red_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")  # Strong red
            strong_green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")  # Strong green
            yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Yellow for neutral
            
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value is not None and cell.value != 'None':
                        val = float(cell.value)
                        if val > 0.5:
                            # Very positive news
                            cell.fill = strong_green_fill
                        elif val > 0:
                            # Positive news
                            cell.fill = green_fill
                        elif val < -0.5:
                            # Very negative news
                            cell.fill = strong_red_fill
                        elif val < 0:
                            # Negative news
                            cell.fill = red_fill
                        else:
                            # Neutral news
                            cell.fill = yellow_fill
                except (ValueError, AttributeError, TypeError):
                    pass
        
        wb.save(output_file)
        print(f"\n✓ Hyperlinks and color coding added to {output_file}")
    except Exception as e:
        print(f"Note: Could not add formatting to Excel: {e}")
    
    print(f"\nOutput saved to {output_file}")
    print(f"Charts saved in 'charts/' folder.")


if __name__ == "__main__":
    process_holdings_file()
    print("\nDone!")
